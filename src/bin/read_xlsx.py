#!/usr/bin/env python3
"""
XLSX Reader Tool for AI Agent Mail TRD
Usage: python3 read_xlsx.py <filename.xlsx>
"""

import sys
import json
import csv
import openpyxl
from openpyxl import load_workbook

def list_sheets(filename):
    """List all available sheets in the workbook"""
    try:
        wb = load_workbook(filename)
        return wb.sheetnames
    except Exception as e:
        print(f"Error listing sheets: {e}")
        return []

def read_xlsx_file(filename, sheet_name=None):
    """Read an Excel file and return its content as formatted text"""
    try:
        # Load the workbook
        wb = load_workbook(filename)

        # Get the specified sheet or active/first sheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                return False
        else:
            sheet = wb.active or wb.worksheets[0]

        print(f"Reading Excel file: {filename}")
        print(f"Sheet name: {sheet.title}")
        print(f"Dimensions: {sheet.dimensions}")
        print("=" * 80)

        # Read all rows
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Skip empty rows
            if not any(row):
                continue

            print(f"Row {row_num}:")
            for col_num, cell_value in enumerate(row, 1):
                if cell_value is not None:
                    print(f"  Column {col_num}: {cell_value}")
            print("-" * 40)

        return True

    except FileNotFoundError:
        print(f"Error: File '{filename}' not found")
        return False
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return False

def convert_xlsx_to_csv(filename, sheet_name=None, output_file=None):
    """Convert Excel sheet to CSV format"""
    try:
        # Load workbook
        wb = load_workbook(filename)

        # Get specified sheet or active/first sheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                return False
        else:
            sheet = wb.active or wb.worksheets[0]

        # If no output file specified, print to stdout
        if not output_file:
            import sys
            csv_writer = csv.writer(sys.stdout)
            for row in sheet.iter_rows(values_only=True):
                if not any(row):
                    continue
                csv_writer.writerow([cell if cell is not None else '' for cell in row])
            return True

        # Write to file
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)

            # Write all rows
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if not any(row):
                    continue
                # Convert None to empty string for CSV
                csv_writer.writerow([cell if cell is not None else '' for cell in row])

        print(f"Converted sheet '{sheet.title}' to {output_file}")
        return output_file

    except Exception as e:
        print(f"Error converting to CSV: {e}")
        return False

def read_xlsx_to_dict(filename, sheet_name=None):
    """Read Excel file and return as list of dictionaries for easier processing"""
    try:
        wb = load_workbook(filename)

        # Get the specified sheet or active/first sheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                return [], []
        else:
            sheet = wb.active or wb.worksheets[0]

        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else f"Column_{len(headers)+1}")

        # Read data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                data.append(row_dict)

        return headers, data

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return [], []

def read_xlsx_to_json(filename, sheet_name=None, indent=2):
    """Read Excel file and return as JSON for easy API integration"""
    try:
        headers, data = read_xlsx_to_dict(filename, sheet_name)

        result = {
            "filename": filename,
            "sheet_name": sheet_name or "default",
            "headers": headers,
            "data": data,
            "row_count": len(data),
            "column_count": len(headers)
        }

        return json.dumps(result, indent=indent, default=str)

    except Exception as e:
        error_result = {
            "error": f"Error converting to JSON: {e}",
            "filename": filename,
            "sheet_name": sheet_name or "default"
        }
        return json.dumps(error_result, indent=indent)

def read_all_sheets_to_json(filename, indent=2):
    """Read all sheets from Excel file and return as JSON"""
    try:
        wb = load_workbook(filename)
        result = {
            "filename": filename,
            "available_sheets": wb.sheetnames,
            "sheets": {}
        }

        for sheet_name in wb.sheetnames:
            headers, data = read_xlsx_to_dict(filename, sheet_name)
            result["sheets"][sheet_name] = {
                "headers": headers,
                "data": data,
                "row_count": len(data),
                "column_count": len(headers)
            }

        return json.dumps(result, indent=indent, default=str)

    except Exception as e:
        error_result = {
            "error": f"Error reading all sheets to JSON: {e}",
            "filename": filename
        }
        return json.dumps(error_result, indent=indent)

def show_help():
    """Show help information"""
    print("XLSX Reader Tool - Read Excel files (.xlsx, .xlsm, .xltx, .xltm)")
    print("")
    print("Usage: read_xlsx.py <filename.xlsx> [options]")
    print("")
    print("Options:")
    print("  --help, -h               Show this help message")
    print("  --sheet <name>           Read specific sheet")
    print("  --json                   Output as JSON")
    print("  --all-sheets-json        Output all sheets as JSON")
    print("  --list-sheets            List all available sheets")
    print("  --csv                    Convert to CSV format")
    print("  --output <file>           Output to file (for any format)")
    print("")
    print("Examples:")
    print("  read_xlsx.py data.xlsx")
    print("  read_xlsx.py data.xlsx --sheet Summary")
    print("  read_xlsx.py data.xlsx --json")
    print("  read_xlsx.py data.xlsx --sheet Summary --json")
    print("  read_xlsx.py data.xlsx --all-sheets-json")
    print("  read_xlsx.py data.xlsx --list-sheets")
    print("  read_xlsx.py data.xlsx --csv")
    print("  read_xlsx.py data.xlsx --csv --output summary.csv")
    print("  read_xlsx.py data.xlsx --json --output data.json")
    print("  read_xlsx.py --help")
    print("")
    print("Features:")
    print("  - Reads Excel (.xlsx, .xlsm, .xltx, .xltm) files")
    print("  - Supports multiple sheets")
    print("  - JSON output for programmatic use")
    print("  - CSV export functionality")
    print("  - Human-readable formatting")
    print("  - Automatic header detection")
    print("  - Empty row handling")

if __name__ == "__main__":
    # Handle --help option without filename requirement
    if len(sys.argv) >= 2 and sys.argv[1] in ["--help", "-h"]:
        show_help()
        sys.exit(0)

    if len(sys.argv) < 2:
        print("Usage: read_xlsx.py <filename.xlsx> [options]")
        print("Use --help for detailed usage information.")
        sys.exit(1)

    filename = sys.argv[1]
    sheet_name = None
    output_json = False
    all_sheets_json = False
    list_sheets_only = False
    csv_output = False
    output_file = None

    # Parse command line arguments
    i = 2
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg in ["--help", "-h"]:
            show_help()
            sys.exit(0)
        elif arg == "--sheet" and i + 1 < len(sys.argv):
            sheet_name = sys.argv[i + 1]
            i += 2
        elif arg == "--json":
            output_json = True
            i += 1
        elif arg == "--all-sheets-json":
            all_sheets_json = True
            i += 1
        elif arg == "--list-sheets":
            list_sheets_only = True
            i += 1
        elif arg == "--csv":
            csv_output = True
            i += 1
        elif arg == "--output" and i + 1 < len(sys.argv):
            output_file = sys.argv[i + 1]
            i += 2
        else:
            print(f"Unknown option: {arg}")
            print("Use --help for usage information.")
            sys.exit(1)

    # Execute based on options
    if list_sheets_only:
        sheets = list_sheets(filename)
        if sheets:
            print(f"Available sheets in {filename}:")
            for i, sheet in enumerate(sheets, 1):
                print(f"  {i}. {sheet}")
        else:
            print("No sheets found or error occurred.")

    elif all_sheets_json:
        print(read_all_sheets_to_json(filename))

    elif csv_output:
        convert_xlsx_to_csv(filename, sheet_name, output_file)

    elif output_json:
        if output_file:
            # Write JSON to file
            with open(output_file, 'w') as f:
                f.write(read_xlsx_to_json(filename, sheet_name))
            print(f"JSON output written to {output_file}")
        else:
            print(read_xlsx_to_json(filename, sheet_name))

    else:
        read_xlsx_file(filename, sheet_name)
